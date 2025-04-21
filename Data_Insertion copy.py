import csv
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange


def fieldInsertion(nameRange, amountRange, fields, ws):
    skipList = []
    skipRow = []
    for row, col in nameRange.cells:
        for key in fields:
            if key not in skipList and row not in skipRow:
                ws.cell(row, col).value = key
                skipList.append(key)
                skipRow.append(row)

    skipList = []
    skipRow = []
    for row, col in amountRange.cells:
        for value in fields.values():
            if value not in skipList and row not in skipRow:
                ws.cell(row, col).value = value
                skipList.append(value)
                skipRow.append(row)




def insertData(bankfile, outputfile):




    income = {
      "Job": 0,
      "Side hustle": 0,
      "Other": 0
    }

    savings = {
        "General Savings": 0
    }

    debts = {
        "Credit Card 1": 0,
        "Credit Card 2": 0,
        "Credit Card 3":0
    }

    bills = {
        "Rent": 0,
        "Insurance 1": 0,
        "Insurance 2": 0,
        "Car Loan": 0,
        "Water Bill": 0,
        "Phone Bill": 0,
        "Internet Bill" : 0,
        "Power Bill" : 0
    }

    variables = {
        "ATM Withdrawals" : 0,
        "Shopping" : 0,
        "Entertainment" : 0,
        "Gas" : 0,
        "Groceries" : 0,
        "Dining" : 0,
        "Other" : 0

    }


    with open(bankfile, mode='r') as csv_file:
        csv_reader = csv.reader(csv_file)
        next(csv_reader, None)
        for row in csv_reader:
            amount = float(row[2])
            transaction_type = row[3]
            typeGroup = row[5]
            description = row[10]
            category = row[11]

            #Income Calculations
            if transaction_type == "Credit" and category == 'Paychecks/Salary':
                income["Job"] += amount

            if transaction_type == "Credit" and description == 'Side hustle Deposit':
                income["Side hustle"] += amount

            if (transaction_type == "Credit" and description != 'Side hustle Deposit' and category != 'Paychecks/Salary') or (category == 'Refunds/Adjustments'):
                income["Other"] += amount

            # Savings Calculations
            if transaction_type == "Debit" and category == 'Savings':
                savings["General Savings"] += amount

            # Debt Calculations
            if transaction_type == "Debit" and category == 'Credit Card Payments' and description == 'Payment to Credit Card 1':
                debts["Credit Card 1"] += amount

            if transaction_type == "Debit" and category == 'Credit Card Payments' and description == 'Payment to Credit Card 2':
                debts["Credit Card 2"] += amount

            if transaction_type == "Debit" and description == 'Transfer To Credit Card 3':
                debts["Credit Card 3"] += amount

            # Bill Calculations
            if transaction_type == "Debit" and typeGroup == 'POS' and 'Rent' in description:
                bills["Rent"] += amount

            if transaction_type == "Debit" and category == 'Insurance' and description == 'Payment to Insurance 1':
                bills["Insurance 1"] += amount

            if transaction_type == "Debit" and category == 'Insurance' and description == 'Payment to Insurance 2':
                bills["Insurance 2"] += amount

            if (transaction_type == "Debit" and category == 'Loans' and description == 'Payment to Car Loan') or (transaction_type == "Debit" and category == 'Online Services' and description == 'Car Loan'):
                bills["Car Loan"] += amount

            if transaction_type == "Debit" and typeGroup == 'POS' and 'water' in description:
                bills["Water Bill"] += amount

            if transaction_type == "Debit" and category == 'Telephone Services' and description == 'Payment to Carrier':
                bills["Phone Bill"] += amount

            if transaction_type == "Debit"  and description == 'Payment to Internet':
                bills["Internet Bill"] += amount

            if transaction_type == "Debit" and category == 'Utilities' and description == 'Payment to Power':
                bills["Power Bill"] += amount

            # Variables
            if transaction_type == "Debit" and typeGroup == 'ATM' and category == 'ATM/Cash Withdrawals':
                variables["ATM Withdrawals"] += amount

            if transaction_type == "Debit" and  (category == 'General Merchandise' or category == 'Clothing/Shoes'):
                variables["Shopping"] += amount

            if transaction_type == "Debit" and category == 'Entertainment':
                variables["Entertainment"] += amount

            if transaction_type == "Debit" and category == 'Gasoline/Fuel':
                variables["Gas"] += amount

            if transaction_type == "Debit" and category == 'Groceries':
                variables["Groceries"] += amount

            if transaction_type == "Debit" and category == 'Restaurants/Dining':
                variables["Dining"] += amount


            if transaction_type == "Debit" and (category == 'Other Expenses' or category == 'Service Charges/Fees' or category == 'Personal Care' or category == 'Securities Trades' or category == 'Healthcare/Medical' or category == 'Automotive Expenses'):
                variables["Other"] += amount



    # Load the Excel workbook using openpyxl
    wb = load_workbook(outputfile)
    ws = wb.active


    # Income insertion
    incomeNameRange = CellRange("E2:E5")
    incomeAmountRange = CellRange("G2:G5")
    fieldInsertion(incomeNameRange, incomeAmountRange, income, ws)

    # Saving insertion
    savingNameRange = CellRange("I9:I16")
    savingAmountRange = CellRange("K9:K16")
    fieldInsertion(savingNameRange, savingAmountRange, savings, ws)

    # Debt insertion
    debtNameRange = CellRange("I20:I28")
    debtAmountRange = CellRange("K20:K28")
    fieldInsertion(debtNameRange, debtAmountRange, debts, ws)

    # Bill insertion
    billNameRange = CellRange("A9:A28")
    billAmountRange = CellRange("C9:C28")
    fieldInsertion(billNameRange, billAmountRange, bills, ws)

    # Variable insertion
    variableNameRange = CellRange("E9:E28")
    variableAmountRange = CellRange("G9:G28")
    fieldInsertion(variableNameRange, variableAmountRange, variables, ws)

    # Save and Export
    wb.save(outputfile)
    wb.close()