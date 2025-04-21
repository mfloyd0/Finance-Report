from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.cell_range import CellRange
from datetime import datetime
from dateutil.relativedelta import relativedelta

from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)

def borderOutline(sheet, borderRange):
    ranges = CellRange(borderRange)
    for row, col in ranges.cells:
        top = Side(style="thick") if (row, col) in ranges.top else None
        left = Side(style="thick") if (row, col) in ranges.left else None
        right = Side(style="thick") if (row, col) in ranges.right else None
        bottom = Side(style="thick") if (row, col) in ranges.bottom else None
        sheet.cell(row, col).border = Border(left, right, top, bottom, outline=True)
    return


def formatFile(file):
    # Get the current date and Subtract one month
    previous_month = datetime.today() - relativedelta(months=1)

    # Format the date to abbr
    previous_month_abr = previous_month.strftime("%b")



    # Load the Excel workbook using openpyxl
    wb = Workbook()
    ws = wb.active

    # Fill color
    lightOrangeFill = PatternFill(start_color='F8CBAD',
                       end_color='F8CBAD',
                       fill_type='solid')

    # Merge and center the cells
    ws.merge_cells('A1:C7')
    # Cell width adjustment
    ws.column_dimensions['A'].width = 19
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 19
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['H'].width = 4
    ws.column_dimensions['I'].width = 19
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 14

    #Budget boxes with field names
    ws.cell(1,1,previous_month_abr).font = Font(size = "100")
    # Bills Format
    ws.cell(8,1,'Bills').fill = lightOrangeFill
    ws.cell(8,2,'Budget').fill = lightOrangeFill
    ws.cell(8,3,'Actual').fill = lightOrangeFill
    ws.cell(29,1,'Total').fill = lightOrangeFill
    # Income Format
    ws.cell(1,5,'Income').fill = lightOrangeFill
    ws.cell(1,6,'Budget').fill = lightOrangeFill
    ws.cell(1,7,'Actual').fill = lightOrangeFill
    ws.cell(6,5,'Total').fill = lightOrangeFill
    # Variable Format
    ws.cell(8,5,'Variable').fill = lightOrangeFill
    ws.cell(8,6,'Budget').fill = lightOrangeFill
    ws.cell(8,7,'Actual').fill = lightOrangeFill
    ws.cell(29,5,'Total').fill = lightOrangeFill
    # Savings Format
    ws.cell(8,9,'Savings').fill = lightOrangeFill
    ws.cell(8,10,'Budget').fill = lightOrangeFill
    ws.cell(8,11,'Actual').fill = lightOrangeFill
    ws.cell(17,9,'Total').fill = lightOrangeFill
    # Debt Format
    ws.cell(19,9,'Debt').fill = lightOrangeFill
    ws.cell(19,10,'Budget').fill = lightOrangeFill
    ws.cell(19,11,'Actual').fill = lightOrangeFill
    ws.cell(29,9,'Total').fill = lightOrangeFill

    # Total Block
    ws.cell(31, 1, 'Bills')
    ws.cell(32, 1, 'Variable')
    ws.cell(33, 1, 'Savings')
    ws.cell(34, 1, 'Debt')
    ws.cell(35, 1, 'Overall Total')

    ws.cell(30, 2, 'Total')


    ##Formulas
    # Income
    ws["G6"] = "=SUM(G2:G5)"
    # Bills
    ws["C29"] = "=SUM(C9:C28)"
    # Variable
    ws["G29"] = "=SUM(G9:G28)"
    # Savings
    ws["K17"] = "=SUM(K9:K16)"
    # Debt
    ws["K29"] = "=SUM(K20:K28)"

    # Total Block
    ws["B31"] = "=C29"
    ws["B32"] = "=G29"
    ws["B33"] = "=K17"
    ws["B34"] = "=K29"
    ws["B35"] = "=SUM(B31:B34)"


    thin = Side(border_style="thin", color="000000")



    billBox = "A8:C29"
    incomeBox = "E1:G6"
    variableBox = "E8:G29"
    savingsBox = "I8:K17"
    debtBox = "I19:K29"

    #Border outline
    borderOutline(ws, billBox)
    borderOutline(ws, incomeBox)
    borderOutline(ws, variableBox)
    borderOutline(ws, savingsBox)
    borderOutline(ws, debtBox)


    # Pie chart
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=31, max_row=34)
    data = Reference(ws, min_col=2, min_row=30, max_row=34)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Budget"
    pie.height = 10
    ws.add_chart(pie, "M8")



    # Save the changes
    wb.save(file)
    wb.close()


